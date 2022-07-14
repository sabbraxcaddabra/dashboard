import dash
from dash import dcc
from dash import Input, Output, callback
from dash import html
import datetime
import pandas as pd
import io

import dash_bootstrap_components as dbc

import plotly.express as px
import plotly.graph_objects as go

import os
import json

import numpy as np

import openpyxl

from . import data_loader

DATA_LOADER = data_loader.DataLoader()
DATA_LOADER.load_data()

HEADER = [
    {'name': ('Направление подготовки, специальность, магистерская программа', 'Код'), 'id': 'spec_code'},
    {'name': ('Направление подготовки, специальность, магистерская программа', 'Название'), 'id': 'spec_name'},
    {'name': ('Количество принятых заявлений', 'Бюджет'), 'id': 'application_b'},
    {'name': ('Количество принятых заявлений', 'Кол-во заявлений / КЦП'), 'id': 'application_b_kcp_ratio'},
    {'name': ('Количество принятых заявлений', 'Контракт'), 'id': 'application_k'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Бюджет'), 'id': 'orig_b'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Ср.балл'), 'id': 'orig_b_ball'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Контракт'), 'id': 'orig_k'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Ср.балл '), 'id': 'orig_k_ball'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Основные места'), 'id': 'orig_osn'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Целевая квота'), 'id': 'orig_celo'},
]

BAC_SPEC_HEADER = HEADER + [
    {'name': ('Согласий при наличии оригинала (договора)', 'Особая квота'), 'id': 'orig_os'},
    {'name': ('Согласий при наличии оригинала (договора)', 'Специальная квота'), 'id': 'orig_spec'},
]

HERE = os.path.dirname(__file__)

DATA_FILE = os.path.abspath(os.path.join(HERE, "..", "data", "stats.xlsx"))
KCP_FILE = os.path.abspath(os.path.join(HERE, "..", "data", "kcp.json"))
TOTAL_KCP_FILE = os.path.abspath(os.path.join(HERE, "..", "data", "total_kcp.json"))

df = pd.read_excel(DATA_FILE)

real_df = DATA_LOADER.load_data

DEFAULT_DICT = {'spec_code': '00.00.00',
   'spec_name': '-',
   'kcp_k': 4,
   'kcp_os': 2,
   'kcp_spec': 2,
   'kcp_cel': 4,
   'kcp_osn': 12
}

with open(KCP_FILE, encoding='utf8') as kcp_file, open(TOTAL_KCP_FILE, encoding='utf8') as total_kcp_file:

    KCP_DICT = json.load(kcp_file)
    TOTAL_KCP_DICT: dict = json.load(total_kcp_file)


def size_merged(sheet, first_letter, last_letter):
    firt_cell_val = str(sheet[f'{first_letter}1'].value)
    firt_cell_len = len(firt_cell_val)
    sum_len = 0
    cells_to_merge = sheet[f'{first_letter}1:{last_letter}1'][0]
    for cell in cells_to_merge:
        col_letter = cell.column_letter
        sum_len += sheet.column_dimensions[col_letter].width

    if 1.3 * firt_cell_len > sum_len:
        firt_cell_len *= 1.3
        dlen = (firt_cell_len - sum_len) / len(cells_to_merge)

        for cell in cells_to_merge:
            col_letter = cell.column_letter
            sheet.column_dimensions[col_letter].width += dlen


def autofit_columns(edu_level, edu_form):

    filepath = os.path.abspath(os.path.join(HERE, "..", f"{edu_level}.xlsx"))

    wb = openpyxl.load_workbook(filepath)
    sheet = wb[edu_form]

    merged_cells = {
        'Бакалавриат': (('B', 'C'), ('D', 'E'), ('F', 'M')),
        'Специалитет': (('B', 'C'), ('D', 'E'), ('F', 'M')),
        'Магистратура': (('B', 'C'), ('D', 'E'), ('F', 'K'))
    }

    merged_cells = merged_cells[edu_level]

    for merged_cell in merged_cells:
        sheet.unmerge_cells(f'{merged_cell[0]}1:{merged_cell[1]}1')

    dims = {}

    rows = list(sheet.rows)
    for row in rows[1:]:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) * 1.3))

    for col, value in dims.items():
        sheet.column_dimensions[col].width = value

    for merged_cell in merged_cells:
        size_merged(sheet, merged_cell[0], merged_cell[1])

    for merged_cell in merged_cells:
        sheet.merge_cells(f'{merged_cell[0]}1:{merged_cell[1]}1')

    wb.save(filepath)


def get_kcp_dict_by_edu_level(edu_level): # Словарь на уровень образование(внутри ключи - формы обучения, значения - словарь со специальностями)
    return DATA_LOADER.total_kcp_dict[edu_level]

def get_kcp_dict_by_edu_form(tmp_kcp_dict, edu_form): # Словарь на форму обуения для словаря, оставшегося при вызове предыдущей функции
    return tmp_kcp_dict[edu_form]

def get_all_edu_forms(edu_level): # Все доступные формы обучения по уровню образования
    return list(DATA_LOADER.total_kcp_dict[edu_level].keys())

def get_all_specs(edu_level): # Все доступные специальности по уровню образования
    return ['Все'] + list(DATA_LOADER.total_kcp_dict[edu_level]['Очное'].keys())


control_elements = html.Div(children=[
    dcc.Interval(id='load_data_interval', interval=300e3),
    html.Button(id='download_all', children='Сформировать полный отчет в Excel'),
    html.Br(),
    dcc.Download(id='mag'),
    dcc.Download(id='bac'),
    dcc.Download(id='spec'),
    dbc.Row(children=[
        dbc.Col(children=[
            html.H3('Уровень образования'),
            dcc.Dropdown(id='edu_level', options=['Бакалавриат', 'Специалитет', 'Магистратура'], value='Бакалавриат', clearable=False)
        ]),
        dbc.Col(children=[
            html.H3('Форма обучения'),
            dcc.Dropdown(id='edu_form', options=['Очное', 'Очно-заочное', 'Заочное'], value='Очное', clearable=False),
        ])
    ])
])

mean_point = html.Div(children=[ # Блок с распределением среднего балла по специальностям
    dbc.Row(children=[
        dbc.Col(children=[
            html.H5('Направление подготовки'),
            dcc.Dropdown(id='spec_names', clearable=False)
        ], width=6)
    ]),
    html.Br(),
    html.Div(id='info_table'),
    html.Div(children=[
        dcc.Graph(id='kvots_plot')
    ], id='kvots_div'),
    html.H3('Распределение по баллам'),
    html.Div('*С учетом выбранных настроек'),
    html.Div('**Для магистратуры указан общий балл'),
    html.H5('Диапазон баллов'),
    dcc.RangeSlider(0, 100, 5, value=[50, 100], id='bal_range'),
    dcc.Graph(id='mean_point_plot')
])

agree_ratio = html.Div(children=[
    html.H3('Соотношение числа оригиналов к общему числу заявлений'),
    html.Div('*С учетом выбранных настроек'),
    dcc.Graph(id='agree_ratio_plot')
])

dop_info = html.Div(children=[
    html.H3('Дополнительная статистика'),
    html.Div('*Не учитывается выбранная форма обучения'),
    html.H3('Распределение по регионам (СПБ и ЛО приведены отдельно)'),
    html.Div(id='spb_lo', style={'marginRight': 700}),
    html.Br(),
    dcc.Graph(id='regio_plot'),
    html.H3('Распределение по гражданству (кроме граждан РФ)'),
    dcc.Graph(id='citiz_plot'),
    html.H3('Соотношение поступающих мужского и женского пола'),
    dcc.Graph(id='gender_plot'),
])

layout = html.Div(children=[
    control_elements,
    mean_point,
    # agree_ratio,
    dop_info
])

@callback(
    [Output('regio_plot', 'figure'), Output('spb_lo', 'children'), Output('gender_plot', 'figure'),
    Output('citiz_plot', 'figure')
    ],
    [Input('load_data_interval', 'n_intervals'), Input('edu_level', 'value'), Input('spec_names', 'value')],
)
def update_data(n, edu_level, spec_name):
    DATA_LOADER.load_data()
    return get_regions_plot(edu_level, spec_name), get_spb_lo(edu_level, spec_name), get_gender_plot(edu_level, spec_name), get_citiz_plot(edu_level, spec_name)

@callback(
    [Output('kvots_plot', 'figure'), Output('kvots_div', 'style')],
    [Input('load_data_interval', 'n_intervals'), Input('edu_level', 'value'), Input('edu_form', 'value'), Input('spec_names', 'value')]
)
def get_kvots_plot(n, edu_level, edu_form, spec_name): # Распределение по формам оплаты
    df = DATA_LOADER.data # Получаем из глобальной области видимости данные

    tmp_df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_edu_form(tmp_df, edu_form)
    tmp_df = get_df_by_spec_name(tmp_df, spec_name)
    counts = pd.value_counts(tmp_df['fintype'])
    tmp_df = pd.DataFrame(data={'Форма оплаты': counts.index, 'Количество': counts.values})

    fig = px.pie(tmp_df, names='Форма оплаты', values='Количество', height=600, title='Распределение количества заявлений по соответсвующим формам оплаты')

    return fig, {'display':'block'}

def get_spec_table_data(tmp_df, spec_name, kcp_dict): # Таблица с данными на 1 специальность

    tmp_df = get_df_by_spec_name(tmp_df, spec_name) # Отбираем по специальности
    applications_b = get_df_by_fintype(tmp_df, 'Бюджет').shape[0] # Кол-во заявлений бюджет
    applications_k = get_df_by_fintype(tmp_df, 'Контракт').shape[0] # Кол-во заявлений контракт
    tmp_df = tmp_df[tmp_df['orig_and_agree'] == 1] # Отбираем только заявлений с согласием и подлинником (или договором)
    # print(tmp_df.loc[:, ['abiturient_id', 'fintype', 'point_mean']])

    budget_points = get_df_by_fintype(tmp_df, 'Бюджет')['point_mean']
    kontract_points = get_df_by_fintype(tmp_df, 'Контракт')['point_mean']
    mean_bal_b = budget_points[budget_points > 0].mean() # Средний балл бюджет
    mean_bal_k = kontract_points[kontract_points > 0.].mean() # Средний балл контракт

    # tmp_df = tmp_df[
    #     tmp_df['orig_and_agree'] == 1]  # Отбираем только заявлений с согласием и подлинником (или договором)

    counts = pd.value_counts(tmp_df['fintype'])

    orig_all_k = counts.get('С оплатой обучения', 0) # Кол-во подлинников контракт
    orig_osn = counts.get('Основные места', 0) # Подлинников на основные места
    orig_celo = counts.get('Целевая квота', 0) # Подлинников на целевую квоту
    orig_os = counts.get('Особая квота', 0) # Подлинников на особую квоту
    orig_spec = counts.get('Специальная квота', 0) # Подлинников на специальную квоту
    orig_all_b = orig_osn + orig_celo + orig_os + orig_spec  # Кол-во подлинников бюджет
    application_b_kcp_ratio = round(applications_b / kcp_dict["kcp_b_all"], 2) if kcp_dict["kcp_b_all"] != 0 else None


    spec_dict = {
        'spec_code': kcp_dict['spec_code'],
        'spec_name': kcp_dict['spec_name'],
        'application_b': applications_b,
        'application_k': applications_k,
        'application_b_kcp_ratio': application_b_kcp_ratio,
        'orig_b': f'{orig_all_b} / {kcp_dict["kcp_b_all"]}',
        'orig_b_ball': round(mean_bal_b, 1),
        'orig_k': f'{orig_all_k} / {kcp_dict["kcp_k_all"]}',
        'orig_k_ball': round(mean_bal_k, 1),
        'orig_osn': f'{orig_osn} / {kcp_dict["kcp_b_all"] - kcp_dict.get("kcp_celo", 0) - kcp_dict.get("kcp_os", 0) - kcp_dict.get("kcp_spec", 0)}',
        'orig_celo': f'{orig_celo} / {kcp_dict.get("kcp_celo", 0)}',
        'orig_os': f'{orig_os} / {kcp_dict.get("kcp_os", 0)}',
        'orig_spec': f'{orig_spec} / {kcp_dict.get("kcp_spec", 0)}'
    }
    return spec_dict

@callback(
    [Output('bac', 'data'), Output('spec', 'data'), Output('mag', 'data')],
    [Input("download_all", "n_clicks")], prevent_initial_call=True
)
def download_all(n_clicks): # Формирует и скачивает все эксель файлы с отчетом по текущей выгрузке из базы
    df = DATA_LOADER.data

    for edu_level in ('Бакалавриат', 'Специалитет', 'Магистратура'):
        header = BAC_SPEC_HEADER if edu_level != 'Магистратура' else HEADER
        tmp_df = get_df_by_edu_level(df, edu_level)

        with pd.ExcelWriter(f'{edu_level}.xlsx', engine='openpyxl', mode='w') as writer:
            for edu_form in get_all_edu_forms(edu_level):
                tmp_tmp_df = get_df_by_edu_form(tmp_df, edu_form)
                specs = get_all_specs(edu_level=edu_level)[1:]
                data = []
                for spec in specs:
                    kcp_dict = get_kcp_dict_by_edu_level(edu_level)
                    kcp_dict = get_kcp_dict_by_edu_form(kcp_dict, edu_form)
                    kcp_dict = kcp_dict[spec]
                    data.append(get_spec_table_data(tmp_tmp_df, spec_name=spec, kcp_dict=kcp_dict))
                data_for_data_table = {}
                for head in header:
                    header_name = head['name']
                    header_id = head['id']
                    header_data = [dat_dict[header_id] for dat_dict in data]
                    data_for_data_table[header_name] = header_data
                table_df = pd.DataFrame(
                    data=data_for_data_table
                )
                table_df.to_excel(writer, sheet_name=edu_form)

    for edu_level in ('Бакалавриат', 'Специалитет', 'Магистратура'):
        for edu_form in get_all_edu_forms(edu_level):
            autofit_columns(edu_level, edu_form)

    return dcc.send_file('Бакалавриат.xlsx'), dcc.send_file('Специалитет.xlsx'), dcc.send_file('Магистратура.xlsx')

@callback(
    Output('info_table', 'children'),
    [Input('load_data_interval', 'n_intervals'), Input('edu_level', 'value'), Input('edu_form', 'value'), Input('spec_names', 'value')]
)
def get_info_table(n, edu_level, edu_form, spec_name): # Отрисовывает таблицу с информацией по выбранному уровню образования, форме обучения и названию специальности

    df = DATA_LOADER.data
    # df = df[df['point_mean'] > 0]
    tmp_df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_edu_form(tmp_df, edu_form)

    if edu_level != 'Магистратура':
        header = BAC_SPEC_HEADER
    else:
        header = HEADER

    if spec_name != 'Все':
        kcp_dict = get_kcp_dict_by_edu_level(edu_level)
        kcp_dict = get_kcp_dict_by_edu_form(kcp_dict, edu_form)
        kcp_dict = kcp_dict[spec_name]
        data = [get_spec_table_data(tmp_df, spec_name=spec_name, kcp_dict=kcp_dict)]
    else:
        data = []
        for spec in get_all_specs(edu_level)[1:]:
            kcp_dict = get_kcp_dict_by_edu_level(edu_level)
            kcp_dict = get_kcp_dict_by_edu_form(kcp_dict, edu_form)
            kcp_dict = kcp_dict[spec]
            data.append(get_spec_table_data(tmp_df, spec_name=spec, kcp_dict=kcp_dict))

    return dash.dash_table.DataTable(
        columns=header,
        data=data,
        merge_duplicate_headers=True,
        style_cell={
                    'text_align': 'left'
                    },
    )

@callback(
    [Output('spec_names', 'options'), Output('spec_names', 'value')],
    [Input('edu_level', 'value')]
)
def get_spec_names(edu_level): # Записывает в качестве опций выпадающего списка со специальностями все специальности по данному уровню образования
    specs = get_all_specs(edu_level)
    return specs, 'Все'

@callback(
    [Output('edu_form', 'options'), Output('edu_form', 'value')],
    [Input('edu_level', 'value')]
)
def get_edu_forms(edu_level): # Записывает в качестве опций выпадающего списка с формами обучения все формы обучения по данному уровню образования
    return get_all_edu_forms(edu_level), 'Очное'

def get_df_by_edu_form(tmp_df, edu_form):
    # Фильтруем по признаку Очное / Очно-Заочное / Заочное
    return tmp_df[tmp_df['edu_form'] == edu_form]

def get_df_by_fintype(tmp_df, fintype):
    # Фильтруем по признаку Бюджет / Контракт
    if fintype != 'Контракт':
        return tmp_df[tmp_df['fintype'] != 'С оплатой обучения']
    else:
        return tmp_df[tmp_df['fintype'] == 'С оплатой обучения']

def get_df_by_edu_level(tmp_df, edu_level):
    # Фильтруем по признаку Бакалавриат / Специалитет / Магистратура
    return tmp_df[tmp_df['edu_level'] == edu_level]

def get_df_by_spec_name(tmp_df, spec_name):
    # Фильтруем по специальности
    if spec_name != 'Все':
        return tmp_df[tmp_df['spec_name'] == spec_name]
    else:
        return tmp_df

@callback(
    Output('mean_point_plot', 'figure'),
    [Input('load_data_interval', 'n_intervals'), Input('edu_level', 'value'), Input('edu_form', 'value'), Input('spec_names', 'value'),
     Input('bal_range', 'value')
     ]
)
def update_mean_point_plot(n, edu_level, edu_form, spec_name, bal_range): # Обновляет график с распределением баллов
    df = DATA_LOADER.data
    df = df[df['point_mean'] > 0]
    tmp_df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_edu_form(tmp_df, edu_form)
    tmp_df = get_df_by_spec_name(tmp_df, spec_name)

    tmp_df = tmp_df[(tmp_df['point_mean'] > bal_range[0]) & (tmp_df['point_mean'] < bal_range[1])]

    fig = px.histogram(data_frame=tmp_df, x='point_mean', nbins=25, marginal='box')
    fig.update_layout(bargap=0.1)

    fig.update_layout(
        xaxis_title="Средний балл",
        yaxis_title="Количество заявлений"
    )
    return fig

@callback(
    Output('agree_ratio_plot', 'figure'),
    [Input('load_data_interval', 'n_intervals'), Input('edu_level', 'value'), Input('edu_form', 'value'), Input('spec_names', 'value'),
     Input('bal_range', 'value')
     ]
)
def agree_ratio(n, edu_level, edu_form, spec_name, bal_range): # Обновляет график отношения числа согласных к общему числу заявлений
    df = DATA_LOADER.data
    tmp_df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_edu_form(tmp_df, edu_form)
    tmp_df = get_df_by_spec_name(tmp_df, spec_name)

    tmp_df = tmp_df[(tmp_df['point_mean'] > bal_range[0]) & (tmp_df['point_mean'] < bal_range[1])]

    agree = tmp_df[tmp_df['original'] == 1]

    counts, bins = np.histogram(agree['point_mean'], bins=range(*bal_range, 2))

    bins = 0.5 * (bins[:-1] + bins[1:])

    counts = counts / tmp_df.shape[0] if tmp_df.shape[0] != 0 else counts

    fig = px.bar(x=bins, y=counts, labels={'x': 'Средний балл', 'y': 'Отношение числа согласий к общему числу заявлений'}, height=650)

    return fig

def get_hostel_num(spec_name):
    df = DATA_LOADER.data
    tmp_df = get_df_by_spec_name(df, spec_name)
    tmp_df = tmp_df.drop_duplicates(subset='abiturient_id')
    counts = pd.value_counts(tmp_df['hostel']).get(1, 0)
    return f'Общежитие требуется: {counts} человек'


def get_regions_plot(edu_level, spec_name):
    df = DATA_LOADER.data

    df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_spec_name(df, spec_name)

    tmp_df = tmp_df.drop_duplicates(subset='abiturient_id')

    tmp_df = tmp_df[(tmp_df['region_name'] != 'г. Санкт-Петербург') & (tmp_df['region_name'] != 'Ленинградская обл.') & (tmp_df['region_name'] != '-')]

    counts = pd.value_counts(tmp_df['region_name'])
    index = counts.index
    values = counts.values

    tmp_df = pd.DataFrame(data={'regio': index, 'counts': values})
    tmp_df = tmp_df.sort_values(by='counts', ascending=False, ignore_index=True)


    if tmp_df.shape[0] > 10:
        tmp_df = tmp_df.iloc[:10, :]

    tmp_df = tmp_df.rename(columns={'regio': 'Регион', 'counts': 'Количество поступающих'})

    fig = px.bar(tmp_df, y='Регион', x='Количество поступающих', orientation='h',
                 category_orders={'Регион': tmp_df['Регион'][::-1]}
                 )

    fig.update_layout(
        title='10 Наиболее популярных регионов',
        yaxis_title="Регион",
        xaxis_title="Количество поступающих",
    )

    return fig

def get_spb_lo(edu_level, spec_name):
    df = DATA_LOADER.data

    df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_spec_name(df, spec_name)

    tmp_df = tmp_df.drop_duplicates(subset='abiturient_id')

    tmp_df = tmp_df[(tmp_df['region_name'] == 'г. Санкт-Петербург') | (tmp_df['region_name'] == 'Ленинградская обл.')]

    counts = pd.value_counts(tmp_df['region_name'])

    tmp_df = pd.DataFrame(data={
        'Регион': counts.index,
        'Количество поступающих': counts.values
    })

    table = dash.dash_table.DataTable(
        data=tmp_df.to_dict('records'),
        style_cell={'font_size': '20px',
                    'text_align': 'center'
                    },
    )

    return table

def get_citiz_plot(edu_level, spec_name):
    df = DATA_LOADER.data

    df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_spec_name(df, spec_name)

    tmp_df = tmp_df.drop_duplicates(subset='abiturient_id')
    tmp_df = tmp_df[tmp_df['country_name'] != 'РОССИЯ']

    counts = pd.value_counts(tmp_df['country_name'])
    index = counts.index[::-1]
    values = counts.values[::-1]

    tmp_df = pd.DataFrame(data={'Гражданство': index, 'Количество поступающих': values})

    fig = px.bar(data_frame=tmp_df, y='Гражданство', x='Количество поступающих', orientation='h')

    fig.update_layout(
        yaxis_title='Гражданство',
        xaxis_title="Количество поступающих"
    )

    return fig

def get_gender_plot(edu_level, spec_name):
    df = DATA_LOADER.data

    df = get_df_by_edu_level(df, edu_level)
    tmp_df = get_df_by_spec_name(df, spec_name)

    tmp_df = tmp_df.drop_duplicates(subset='abiturient_id')

    counts = pd.value_counts(tmp_df['gender_id'])
    mens = counts.get(1, 0)
    womens = counts.get(2, 0)

    fig = go.Figure(data=[
        go.Bar(x=['Мужчины'], y=[mens], name='Мужчины'),
        go.Bar(x=['Женщины'], y=[womens], name='Женщины')
    ]
    )

    return fig


